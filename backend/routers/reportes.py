"""
MermaZero — Router de Reportes Excel
Genera archivos .xlsx con formato profesional para:
  - Inventario completo
  - Reporte de mermas
  - Reporte de alertas
"""

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from datetime import date
from io import BytesIO

from backend.database.database import get_db
from backend.models.models import Lote, EstadoLote, NivelAlerta
from backend.models.motor_alertas import evaluar_lote
from backend.routers.auth_utils import obtener_usuario_actual
from backend.models.models import Usuario

router = APIRouter(prefix="/api/reportes", tags=["Reportes"])

# ── PALETA MAKRO ──────────────────────────────────────────────────
AZUL_DARK   = "0A1628"
AZUL_MID    = "0052CC"
AZUL_LIGHT  = "EFF6FF"
AMARILLO    = "F5A800"
AMARILLO_BG = "FFFBEB"
ROJO        = "DC2626"
ROJO_BG     = "FEF2F2"
VERDE       = "059669"
VERDE_BG    = "ECFDF5"
MORADO      = "7C3AED"
MORADO_BG   = "F5F3FF"
GRIS_BG     = "F8FAFC"
GRIS_TXT    = "64748B"
BLANCO      = "FFFFFF"

THIN_BORDER = Border(
    left=Side(style='thin', color='E2E8F0'),
    right=Side(style='thin', color='E2E8F0'),
    top=Side(style='thin', color='E2E8F0'),
    bottom=Side(style='thin', color='E2E8F0')
)


def _excel_response(wb: Workbook, filename: str) -> StreamingResponse:
    """Convierte un Workbook a StreamingResponse descargable."""
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


def _banner(ws, titulo: str, subtitulo: str, ncols: int):
    """Dibuja el banner MermaZero en las primeras filas."""
    col_last = get_column_letter(ncols)

    # Fondo azul oscuro en todo el banner
    for row in range(1, 4):
        for col in range(1, ncols + 1):
            ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=AZUL_DARK)

    # Logo
    ws.merge_cells("A1:C3")
    c = ws["A1"]
    c.value = "MermaZero"
    c.font = Font(size=22, bold=True, color=AMARILLO, name="Calibri")
    c.fill = PatternFill("solid", fgColor=AZUL_DARK)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # Título
    ws.merge_cells(f"D1:{col_last}1")
    c = ws["D1"]
    c.value = titulo
    c.font = Font(size=13, bold=True, color=BLANCO)
    c.fill = PatternFill("solid", fgColor=AZUL_DARK)
    c.alignment = Alignment(horizontal="right", vertical="center", indent=2)

    # Subtítulo
    ws.merge_cells(f"D2:{col_last}2")
    c = ws["D2"]
    c.value = subtitulo
    c.font = Font(size=10, color="8899BB")
    c.fill = PatternFill("solid", fgColor=AZUL_DARK)
    c.alignment = Alignment(horizontal="right", vertical="center", indent=2)

    # Fecha
    ws.merge_cells(f"D3:{col_last}3")
    c = ws["D3"]
    c.value = f"Generado el {date.today().strftime('%d/%m/%Y')}"
    c.font = Font(size=9, italic=True, color=AMARILLO)
    c.fill = PatternFill("solid", fgColor=AZUL_DARK)
    c.alignment = Alignment(horizontal="right", vertical="center", indent=2)

    ws.row_dimensions[1].height = 16
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 18

    # Línea amarilla
    ws.merge_cells(f"A4:{col_last}4")
    ws["A4"].fill = PatternFill("solid", fgColor=AMARILLO)
    ws.row_dimensions[4].height = 4


def _kpi_cards(ws, kpis: list, ncols: int):
    """Dibuja las tarjetas KPI (fila 6-8)."""
    col_last = get_column_letter(ncols)
    kpi_w = max(1, ncols // len(kpis))
    start_col = 1
    for label, val, color, bg in kpis:
        c1 = get_column_letter(start_col)
        c2 = get_column_letter(min(start_col + kpi_w - 1, ncols))
        ws.merge_cells(f"{c1}6:{c2}6")
        cl = ws[f"{c1}6"]
        cl.value = label
        cl.font = Font(size=8, bold=True, color=GRIS_TXT)
        cl.fill = PatternFill("solid", fgColor=bg)
        cl.alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells(f"{c1}7:{c2}8")
        cv = ws[f"{c1}7"]
        cv.value = val
        cv.font = Font(size=16, bold=True, color=color)
        cv.fill = PatternFill("solid", fgColor=bg)
        cv.alignment = Alignment(horizontal="center", vertical="center")
        start_col += kpi_w

    ws.row_dimensions[6].height = 16
    ws.row_dimensions[7].height = 24
    ws.row_dimensions[8].height = 8


def _headers(ws, headers: list, row: int):
    """Dibuja los headers de la tabla."""
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = Font(bold=True, color=BLANCO, size=10.5)
        c.fill = PatternFill("solid", fgColor=AZUL_DARK)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = THIN_BORDER
    ws.row_dimensions[row].height = 24


def _footer(ws, row: int, ncols: int):
    """Dibuja el footer del reporte."""
    col_last = get_column_letter(ncols)
    ws.merge_cells(f"A{row}:{col_last}{row}")
    fc = ws[f"A{row}"]
    fc.value = "MermaZero © 2026 · Desarrollado para Makro Chincha · UPSJB Ingeniería de Sistemas"
    fc.font = Font(size=8, italic=True, color=GRIS_TXT)
    fc.alignment = Alignment(horizontal="center")


def _page_setup(ws):
    """Configura la hoja para impresión horizontal."""
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.3, bottom=0.3)


def _nivel_info(nivel: str):
    """Retorna (bg, txt, icon) según el nivel de alerta."""
    if nivel == "ALERTA_ROJA":
        return ROJO_BG, ROJO, "🔴 ROJA"
    elif nivel == "ALERTA_AMARILLA":
        return AMARILLO_BG, "D97706", "🟡 AMARILLA"
    else:
        return VERDE_BG, VERDE, "✅ NORMAL"


# ══════════════════════════════════════════════════════════════════
# REPORTE 1 — INVENTARIO COMPLETO
# ══════════════════════════════════════════════════════════════════
@router.get("/inventario")
def reporte_inventario(
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(obtener_usuario_actual)
):
    lotes = db.query(Lote).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"
    _page_setup(ws)

    NCOLS = 9
    _banner(ws, "REPORTE DE INVENTARIO", "Makro Chincha · Sistema de Control de Mermas 2026", NCOLS)

    # KPIs
    total     = len(lotes)
    en_stock  = sum(1 for l in lotes if l.estado == EstadoLote.EN_STOCK)
    ev_list   = [evaluar_lote(l.fecha_vencimiento, l.producto) for l in lotes if l.estado == EstadoLote.EN_STOCK]
    rojas     = sum(1 for e in ev_list if e["nivel_alerta"] == NivelAlerta.ALERTA_ROJA)
    amarillas = sum(1 for e in ev_list if e["nivel_alerta"] == NivelAlerta.ALERTA_AMARILLA)
    merma     = sum((l.cantidad or 0) * (l.costo_unitario or 0)
                    for l in lotes if l.estado in (EstadoLote.VENCIDO, EstadoLote.DADO_DE_BAJA))

    _kpi_cards(ws, [
        ("TOTAL LOTES",      str(total),         AZUL_MID, AZUL_LIGHT),
        ("EN STOCK",         str(en_stock),       VERDE,    VERDE_BG),
        ("ALERTA ROJA",      str(rojas),          ROJO,     ROJO_BG),
        ("ALERTA AMARILLA",  str(amarillas),      "D97706", AMARILLO_BG),
        ("MERMA VALORIZADA", f"S/. {merma:.2f}",  MORADO,   MORADO_BG),
    ], NCOLS)

    HEADER_ROW = 10
    _headers(ws, ["Código", "Producto", "Categoría", "Cantidad", "Proveedor",
                  "Vencimiento", "Días", "Estado", "Alerta"], HEADER_ROW)

    for i, lote in enumerate(lotes, start=HEADER_ROW + 1):
        ev = evaluar_lote(lote.fecha_vencimiento, lote.producto) if lote.estado == EstadoLote.EN_STOCK else {}
        nivel = ev.get("nivel_alerta", NivelAlerta.SIN_ALERTA).value if hasattr(ev.get("nivel_alerta", ""), "value") else str(ev.get("nivel_alerta", "SIN_ALERTA"))
        dias  = ev.get("dias_para_vencer", None)
        bg, txt, icon = _nivel_info(nivel)

        row_data = [
            lote.codigo, lote.producto, lote.categoria,
            f"{lote.cantidad} {lote.unidad}", lote.proveedor,
            lote.fecha_vencimiento.strftime("%d/%m/%Y") if lote.fecha_vencimiento else "–",
            dias if dias is not None else "–",
            lote.estado.value.replace("_", " "),
            icon,
        ]
        for c_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=c_idx, value=val)
            cell.font = Font(size=10, color="1E293B")
            cell.alignment = Alignment(
                horizontal="center" if c_idx in (1, 4, 6, 7, 9) else "left",
                vertical="center", indent=1 if c_idx == 2 else 0
            )
            cell.border = THIN_BORDER
            if (i - HEADER_ROW) % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=GRIS_BG)
            if c_idx == 7 and dias is not None:
                cell.font = Font(bold=True, size=12, color=txt)
                cell.fill = PatternFill("solid", fgColor=bg)
            if c_idx == 9:
                cell.fill = PatternFill("solid", fgColor=bg)
                cell.font = Font(bold=True, size=9.5, color=txt)
        ws.row_dimensions[i].height = 22

    _footer(ws, HEADER_ROW + len(lotes) + 2, NCOLS)

    for i, w in enumerate([10, 28, 19, 13, 22, 13, 7, 11, 13], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{HEADER_ROW + 1}"

    return _excel_response(wb, f"MermaZero_Inventario_{date.today()}.xlsx")


# ══════════════════════════════════════════════════════════════════
# REPORTE 2 — MERMAS
# ══════════════════════════════════════════════════════════════════
@router.get("/mermas")
def reporte_mermas(
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(obtener_usuario_actual)
):
    mermas = db.query(Lote).filter(
        Lote.estado.in_([EstadoLote.VENCIDO, EstadoLote.DADO_DE_BAJA])
    ).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Mermas"
    _page_setup(ws)

    NCOLS = 7
    _banner(ws, "REPORTE DE MERMAS", "Makro Chincha · Productos dados de baja por vencimiento 2026", NCOLS)

    total_merma = sum((l.cantidad or 0) * (l.costo_unitario or 0) for l in mermas)
    _kpi_cards(ws, [
        ("TOTAL MERMAS",    str(len(mermas)),          ROJO,   ROJO_BG),
        ("TOTAL VALORIZADO", f"S/. {total_merma:.2f}", MORADO, MORADO_BG),
        ("VENCIDOS",        str(sum(1 for l in mermas if l.estado == EstadoLote.VENCIDO)), "D97706", AMARILLO_BG),
        ("DADOS DE BAJA",   str(sum(1 for l in mermas if l.estado == EstadoLote.DADO_DE_BAJA)), GRIS_TXT, GRIS_BG),
    ], NCOLS)

    HEADER_ROW = 10
    _headers(ws, ["Código", "Producto", "Categoría", "Cantidad",
                  "Vencimiento", "Estado", "Merma (S/.)"], HEADER_ROW)

    for i, lote in enumerate(mermas, start=HEADER_ROW + 1):
        merma_val = (lote.cantidad or 0) * (lote.costo_unitario or 0)
        row_data = [
            lote.codigo, lote.producto, lote.categoria,
            f"{lote.cantidad} {lote.unidad}",
            lote.fecha_vencimiento.strftime("%d/%m/%Y") if lote.fecha_vencimiento else "–",
            lote.estado.value.replace("_", " "),
            round(merma_val, 2),
        ]
        for c_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=c_idx, value=val)
            cell.font = Font(size=10, color="1E293B")
            cell.alignment = Alignment(
                horizontal="center" if c_idx in (1, 4, 5, 7) else "left",
                vertical="center", indent=1 if c_idx == 2 else 0
            )
            cell.border = THIN_BORDER
            if (i - HEADER_ROW) % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=GRIS_BG)
            if c_idx == 7:
                cell.font = Font(bold=True, size=11, color=MORADO)
                cell.fill = PatternFill("solid", fgColor=MORADO_BG)
        ws.row_dimensions[i].height = 22

    # Fila TOTAL
    total_row = HEADER_ROW + len(mermas) + 1
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, color=BLANCO)
    ws.cell(row=total_row, column=1).fill = PatternFill("solid", fgColor=AZUL_DARK)
    ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="center")
    ws.merge_cells(f"B{total_row}:F{total_row}")
    ws.cell(row=total_row, column=2, value="Merma total valorizada").font = Font(bold=True, color=BLANCO)
    ws.cell(row=total_row, column=2).fill = PatternFill("solid", fgColor=AZUL_DARK)
    c_total = ws.cell(row=total_row, column=7, value=round(total_merma, 2))
    c_total.font = Font(bold=True, size=13, color=BLANCO)
    c_total.fill = PatternFill("solid", fgColor=MORADO)
    c_total.alignment = Alignment(horizontal="center")
    ws.row_dimensions[total_row].height = 26

    _footer(ws, total_row + 2, NCOLS)

    for i, w in enumerate([10, 28, 19, 14, 14, 14, 14], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{HEADER_ROW + 1}"

    return _excel_response(wb, f"MermaZero_Mermas_{date.today()}.xlsx")


# ══════════════════════════════════════════════════════════════════
# REPORTE 3 — ALERTAS ACTIVAS
# ══════════════════════════════════════════════════════════════════
@router.get("/alertas")
def reporte_alertas(
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(obtener_usuario_actual)
):
    lotes_stock = db.query(Lote).filter(Lote.estado == EstadoLote.EN_STOCK).all()
    alertas = []
    for lote in lotes_stock:
        ev = evaluar_lote(lote.fecha_vencimiento, lote.producto)
        if ev["nivel_alerta"] != NivelAlerta.SIN_ALERTA:
            alertas.append({
                "lote":    lote,
                "ev":      ev,
                "nivel":   ev["nivel_alerta"].value if hasattr(ev["nivel_alerta"], "value") else str(ev["nivel_alerta"]),
                "dias":    ev["dias_para_vencer"],
                "mensaje": ev["mensaje"],
            })
    alertas.sort(key=lambda x: x["dias"])

    wb = Workbook()
    ws = wb.active
    ws.title = "Alertas"
    _page_setup(ws)

    NCOLS = 6
    _banner(ws, "REPORTE DE ALERTAS ACTIVAS", "Makro Chincha · Productos que requieren atención inmediata", NCOLS)

    rojas    = sum(1 for a in alertas if "ROJA" in a["nivel"])
    amarillas= sum(1 for a in alertas if "AMARILLA" in a["nivel"])
    _kpi_cards(ws, [
        ("TOTAL ALERTAS",   str(len(alertas)), ROJO,   ROJO_BG),
        ("ALERTA ROJA",     str(rojas),        ROJO,   ROJO_BG),
        ("ALERTA AMARILLA", str(amarillas),    "D97706", AMARILLO_BG),
    ], NCOLS)

    HEADER_ROW = 10
    _headers(ws, ["Código", "Producto", "Categoría", "Días para vencer", "Nivel", "Mensaje"], HEADER_ROW)

    for i, a in enumerate(alertas, start=HEADER_ROW + 1):
        bg, txt, icon = _nivel_info(a["nivel"])
        row_data = [
            a["lote"].codigo,
            a["lote"].producto,
            a["lote"].categoria,
            a["dias"],
            icon,
            a["mensaje"],
        ]
        for c_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=c_idx, value=val)
            cell.font = Font(size=10, color="1E293B")
            cell.alignment = Alignment(
                horizontal="center" if c_idx in (1, 4, 5) else "left",
                vertical="center", wrap_text=(c_idx == 6)
            )
            cell.border = THIN_BORDER
            if (i - HEADER_ROW) % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=GRIS_BG)
            if c_idx == 4:
                cell.font = Font(bold=True, size=12, color=txt)
                cell.fill = PatternFill("solid", fgColor=bg)
            if c_idx == 5:
                cell.fill = PatternFill("solid", fgColor=bg)
                cell.font = Font(bold=True, size=9.5, color=txt)
        ws.row_dimensions[i].height = 26

    _footer(ws, HEADER_ROW + len(alertas) + 2, NCOLS)

    for i, w in enumerate([10, 28, 19, 16, 13, 45], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{HEADER_ROW + 1}"

    return _excel_response(wb, f"MermaZero_Alertas_{date.today()}.xlsx")
